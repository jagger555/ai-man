from __future__ import annotations

import os
from collections import Counter
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dependency may not be installed yet
    load_workbook = None


class VisitorAnalyticsService:
    def __init__(self, path: str | Path | None = None):
        self._path = Path(path) if path else _default_analytics_path()

    def get_summary(self) -> dict[str, object]:
        return dict(_load_summary(str(self._path)))


@lru_cache(maxsize=4)
def _load_summary(path_text: str) -> dict[str, object]:
    path = Path(path_text)
    if load_workbook is None or not path.exists():
        return {}

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        indexes = {header: index for index, header in enumerate(headers)}

        total_visits = 0
        tourist_ids: set[str] = set()
        attraction_counts: Counter[str] = Counter()
        gender_counts: Counter[str] = Counter()
        age_group_counts: Counter[str] = Counter()
        month_counts: Counter[str] = Counter()
        stay_total = 0.0
        total_cost = 0.0
        satisfaction_total = 0.0
        group_size_total = 0.0

        for row in rows:
            total_visits += 1
            tourist_id = _cell(row, indexes, "tourist_id")
            if tourist_id:
                tourist_ids.add(str(tourist_id))
            attraction = str(_cell(row, indexes, "attraction_name") or "未知景点")
            attraction_counts[attraction] += 1
            gender = str(_cell(row, indexes, "gender") or "未知")
            gender_counts[gender] += 1

            age = _to_float(_cell(row, indexes, "age"))
            age_group_counts[_age_group(age)] += 1
            visit_date = _cell(row, indexes, "visit_date")
            month_label = _month_label(visit_date)
            if month_label:
                month_counts[month_label] += 1

            stay_total += _to_float(_cell(row, indexes, "stay_duration"))
            total_cost += _to_float(_cell(row, indexes, "total_cost"))
            satisfaction_total += _to_float(_cell(row, indexes, "satisfaction"))
            group_size_total += _to_float(_cell(row, indexes, "group_size"))

        if total_visits == 0:
            return {}

        return {
            "source_file": path.name,
            "sheet_name": sheet.title,
            "total_visits": total_visits,
            "unique_tourists": len(tourist_ids),
            "average_stay_duration": round(stay_total / total_visits, 1),
            "average_total_cost": round(total_cost / total_visits, 2),
            "average_satisfaction": round(satisfaction_total / total_visits, 2),
            "average_group_size": round(group_size_total / total_visits, 1),
            "peak_month": _top_label(month_counts),
            "top_attractions": _top_items(attraction_counts, 5),
            "gender_distribution": _top_items(gender_counts, 4),
            "age_groups": _top_items(age_group_counts, 5),
        }
    except Exception:
        return {}


def _default_analytics_path() -> Path:
    default_path = (
        Path(__file__).resolve().parents[3]
        / "示范景区公开资料包"
        / "景点景区旅游数据行为分析数据.xlsx"
    )
    return Path(os.getenv("VISITOR_ANALYTICS_XLSX_PATH", str(default_path)))


def _cell(row: tuple[Any, ...], indexes: dict[str, int], key: str) -> Any:
    index = indexes.get(key)
    if index is None or index >= len(row):
        return None
    return row[index]


def _to_float(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value))
    except (TypeError, ValueError):
        return 0.0


def _age_group(age: float) -> str:
    if age <= 0:
        return "未知"
    if age < 18:
        return "18岁以下"
    if age < 30:
        return "18-29岁"
    if age < 45:
        return "30-44岁"
    if age < 60:
        return "45-59岁"
    return "60岁以上"


def _month_label(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    text = str(value or "")
    return text[:7] if len(text) >= 7 else ""


def _top_items(counter: Counter[str], limit: int) -> list[dict[str, object]]:
    total = sum(counter.values())
    return [
        {
            "label": label,
            "count": count,
            "share": round(count / total, 3) if total else 0.0,
        }
        for label, count in counter.most_common(limit)
    ]


def _top_label(counter: Counter[str]) -> str:
    return counter.most_common(1)[0][0] if counter else ""
